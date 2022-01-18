from openpyxl import load_workbook

infile = load_workbook("toexcel.xlsx")
product_list=infile("sheet1")

products_per_supplier = {}

for product_row in range(2, products_per_supplier.max_row + 1):
  supplier_name=product_list.cell(product_row, 4).value
  
  #calculation for products per supplier
  if supplier_name in products_per_supplier:
    current_nu_products = products_per_supplier[supplier_name] 
    products_per_supplier[supplier_name] = current_nu_products + 1
  else:
    products_per_supplier[supplier_name] = 1  
print(products_per_supplier) 